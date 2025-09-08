import requests
import json
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers
from datetime import datetime, timedelta

def generate_report(target_project_name, output_stream):
    """
    Generate an FSP Excel report filtered by project name.

    Args:
        target_project_name (str): The project name to filter payments.
        output_stream (file-like): A BytesIO or file-like object to write the Excel file to.
    """

    # === Load config ===
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, 'system_config.json')
    template_path = os.path.join(script_dir, 'Template.xlsx')

    with open(config_path, 'r', encoding='utf-8') as f:
        config = json.load(f)

    # === Bitrix settings ===
    bitrix_webhook = config["B24_WEBHOOK_URL"]
    payment_entity_type_id = config["PAYMENT_ENTITY_TYPE_ID"]

    # === Fields from config ===
    fields = config["EXCEL_EXPORT_FIELDS"]
    FIRST_NAME_FIELD = fields["FIRST_NAME_FIELD"]
    SURNAME_FIELD = fields["SURNAME_FIELD"]
    PATRONYMIC_FIELD = fields["PATRONYMIC_FIELD"]
    PAYMENT_AMOUNT_FIELD = fields["PAYMENT_AMOUNT_FIELD"]
    ID_NUMBER_FIELD = fields["ID_NUMBER_FIELD"]
    REGION_FIELD = fields["REGION_FIELD"]
    PROJECT_NAME_FIELD = fields["PROJECT_NAME_FIELD"]

    print("📡 Fetching payments from Bitrix...")
    print(f"🔍 Filtering by project: '{target_project_name}'")

    # === Fetch all payment records ===
    start = 0
    all_payments = []

    while True:
        url = f"{bitrix_webhook}/crm.item.list"
        payload = {
            "entityTypeId": payment_entity_type_id,
            "start": start
        }

        response = requests.post(url, json=payload)
        if response.status_code != 200:
            raise Exception(f"❌ Request failed: {response.status_code} {response.text}")

        data = response.json()
        items = data.get("result", {}).get("items", [])
        if not items:
            break

        all_payments.extend(items)
        start += 50

    # === Filter payments ===
    filtered = [
        p for p in all_payments
        if p.get(PROJECT_NAME_FIELD, "").strip() == target_project_name
    ]

    print(f"\n📦 Found {len(filtered)} matching payments.\n")

    # === Build rows ===
    rows = []
    for idx, p in enumerate(filtered, start=1):
        rows.append([
            idx,                               # A: Index
            p.get(SURNAME_FIELD, ""),          # B: Фамилия получателя
            p.get(FIRST_NAME_FIELD, ""),       # C: Имя получателя
            p.get(PATRONYMIC_FIELD, ""),       # D: Отчество получателя
            str(p.get(ID_NUMBER_FIELD, "")),   # E: ИНН получателя (text)
            p.get(PAYMENT_AMOUNT_FIELD, ""),   # F: Сумма выплаты
            "СОМ",                             # G: Валюта
            p.get(REGION_FIELD, "")            # H: Город выплаты
        ])

    # === Load template ===
    if not os.path.exists(template_path):
        raise FileNotFoundError("❌ Template.xlsx not found.")

    wb = load_workbook(template_path)
    ws = wb.active

    start_row = 10  # A10
    start_col = 1   # column A

    # --- Find the "Итого:" cell location ---
    itogo_cell = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == "Итого:":
                itogo_cell = cell
                break
        if itogo_cell:
            break

    if not itogo_cell:
        print("⚠️ Could not find a cell containing 'Итого:'. Skipping spacing/sum logic.")
        for r_i, row_data in enumerate(rows):
            for c_i, value in enumerate(row_data):
                cell = ws.cell(row=start_row + r_i, column=start_col + c_i, value=value)
                if c_i == 4:
                    cell.number_format = numbers.FORMAT_TEXT
        wb.save(output_stream)
        return

    itogo_row = itogo_cell.row
    itogo_col = itogo_cell.column

    # --- Ensure enough space: insert rows above 'Итого:' if needed ---
    capacity = itogo_row - start_row
    needed = len(rows)
    extra = needed - capacity
    if extra > 0:
        ws.insert_rows(itogo_row, amount=extra)
        itogo_row += extra

    # --- Write data rows ---
    for r_i, row_data in enumerate(rows):
        for c_i, value in enumerate(row_data):
            cell = ws.cell(row=start_row + r_i, column=start_col + c_i, value=value)
            if c_i == 4:  # Column E (ID)
                cell.number_format = numbers.FORMAT_TEXT

    last_data_row = start_row + max(len(rows), 0) - 1 if rows else (start_row - 1)

    # --- Delete empty rows between last data row and 'Итого:' ---
    delete_start = last_data_row + 1
    delete_end = itogo_row - 1
    if delete_end >= delete_start:
        amount = delete_end - delete_start + 1
        ws.delete_rows(delete_start, amount)
        itogo_row -= amount
        print(f"🧹 Deleted {amount} empty row(s) between data and 'Итого:'.")

    # --- Put SUM formula ---
    sum_cell = ws.cell(row=itogo_row, column=itogo_col + 5)
    if last_data_row >= start_row:
        formula = f"=SUM(F{start_row}:F{last_data_row})"
        sum_cell.value = formula
        ws["C7"] = formula
    else:
        sum_cell.value = "=0"
        ws["C7"] = "=0"

    # --- Fill header info ---
    today = datetime.today()
    tomorrow = today + timedelta(days=4)

    ws.merge_cells("C1:F1")
    ws["C1"] = target_project_name
    ws["C3"] = tomorrow.strftime("%Y-%m-%d")
    ws["C6"] = today.strftime("%Y-%m-%d")

    # Save to provided stream
    wb.save(output_stream)

if __name__ == "__main__":
    import io
    project = input("Enter project name: ").strip()
    with open("Output.xlsx", "wb") as f:
        generate_report(project, f)
    print("✅ Report generated: Output.xlsx")
